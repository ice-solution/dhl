#!/usr/bin/env python3
"""Extract field entitlements from fields_details.xlsx OVERVIEW FORM SUMMARY"""
import json
import re
import openpyxl

wb = openpyxl.load_workbook('fields_details.xlsx', data_only=True)
ws = wb['OVERVIEW FORM SUMMARY']

categories = []
for c in range(6, 19):
    val = ws.cell(2, c).value
    if val and str(val).strip():
        name = re.sub(r'\s+', ' ', str(val).replace('\n', ' ')).strip()
        categories.append({'col': c, 'key': name, 'label': name})

def is_visible(cell_val):
    if cell_val is None:
        return False
    s = str(cell_val).strip()
    if not s or s in ('✕', '×', 'X', 'x'):
        return False
    return '✓' in s or 'INCLUDED' in s.upper()

fields = []
current_section = ''

for r in range(2, 81):
    a = ws.cell(r, 1).value
    b = ws.cell(r, 2).value
    c = ws.cell(r, 3).value
    d = ws.cell(r, 4).value

    if a and str(a).strip():
        current_section = str(a).strip()

    if not b or not str(b).strip():
        continue
    field_name = str(b).strip()
    if field_name.upper() in ('YES', 'NO'):
        continue
    if 'CATEGORY' in field_name.upper() and 'dropdown' in field_name.lower():
        field_key = 'category_dropdown_list'
    else:
        field_key = re.sub(r'[^a-z0-9]+', '_', field_name.lower()).strip('_')

    visibility = {cat['key']: is_visible(ws.cell(r, cat['col']).value) for cat in categories}
    fields.append({
        'section': current_section,
        'name': field_name,
        'key': field_key,
        'typeHint': str(c or '').strip(),
        'optionsHint': str(d or '').strip() if d else '',
        'visibility': visibility,
    })

out = {'categories': categories, 'fields': fields}
with open('data/field-entitlements.json', 'w', encoding='utf-8') as fp:
    json.dump(out, fp, indent=2, ensure_ascii=False)

print(f'Wrote {len(fields)} fields for {len(categories)} categories')
